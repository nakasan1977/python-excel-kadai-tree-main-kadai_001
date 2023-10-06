import openpyxl
import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws["b4"] = "請求書"
ws["b5"] = "株式会社ABC"  # type: ignore
ws["b6"] = "TEL:03-1234-5678 FAX:03-1234-5678"  # type: ignore
ws["b7"] = "担当者名:鈴木一郎 様"  # type: ignore

ws["f4"] = "No"  # type: ignore
ws["f5"] = "日付"  # type: ignore

today = datetime.datetime.now()
print(today.strftime("%Y/%m/%d"))
ws["g5"] = today.strftime("%Y/%m/%d")
ws["b10"] = "商品名"
ws["c10"] = "数量"
ws["d10"] = "単価"
ws["e10"] = "金額"

products = [{"name": "商品A", "quantity": 2, "unite_price": 10000,},{"name": "商品B", "quantity": 1, "unite_price": 15000, },{"name": "商品C", "quantity": 2, "unite_price": 20000, }]

current_row = ws.max_row + 1
for product in products:
    ws[f"B{current_row}"] = product["name"]
    ws[f"C{current_row}"] = product["quantity"]
    ws[f"D{current_row}"] = product["unite_price"]
    current_row = current_row + 1

for i in range(11, 16):
  if ws[f"C{i}"].value is None:
     ws[f"e{i}"] = ""
  else:
     ws[f"E{i}"] =f"=C{i} * D{i}"
  i= i+1


# E列の合計を計算
sum_formula = f"=SUM(E11:E{current_row - 1})"
ws[f"E{current_row}"] = sum_formula

ws["b15"] = "小計"
ws["b16"] = "消費税"
ws["b17"] = "合計"

ws["e15"] = sum_formula
ws["e16"] = "=E15*0.1"
ws["e17"] = "=SUM(E15:E16)"

# ファイルを保存
wb.save("請求書_20231003.xlsx")
